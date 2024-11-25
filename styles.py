from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Colors
HEADER_GREEN = "90EE90"
LIGHT_GREEN = "D0F0C0"

# Styles
header_fill = PatternFill(start_color=HEADER_GREEN, end_color=HEADER_GREEN, fill_type="solid")
subheader_fill = PatternFill(start_color=LIGHT_GREEN, end_color=LIGHT_GREEN, fill_type="solid")

header_font = Font(name='Arial', size=16, bold=True)
normal_font = Font(name='Arial', size=11)
bold_font = Font(name='Arial', size=11, bold=True)

center_align = Alignment(horizontal='center', vertical='center')
left_align = Alignment(horizontal='left', vertical='center')
right_align = Alignment(horizontal='right', vertical='center')

border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)
