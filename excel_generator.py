import openpyxl
from openpyxl.utils import get_column_letter
from styles import *
from datetime import datetime

class InvoiceGenerator:
    def __init__(self):
        self.wb = openpyxl.Workbook()
        self.ws = self.wb.active
        self.setup_page_layout()

    def setup_page_layout(self):
        # Set column widths
        self.ws.column_dimensions['A'].width = 15
        self.ws.column_dimensions['B'].width = 30
        self.ws.column_dimensions['C'].width = 15
        self.ws.column_dimensions['D'].width = 15

        # Company header
        self.ws.merge_cells('A1:D2')
        header_cell = self.ws['A1']
        header_cell.value = "ARNAV SERVICES"
        header_cell.font = header_font
        header_cell.fill = header_fill
        header_cell.alignment = center_align

        # Address section - excluding D3 and D4 for PAN and phone number
        address_cells = ['A3:C3', 'A4:C4', 'A5:D5', 'A6:D6']
        for cell_range in address_cells:
            self.ws.merge_cells(cell_range)

    def add_company_details(self, details):
        company_rows = [
            details['name'],
            details['address_line1'],
            details['address_line2'],
            f"{details['city']} {details['state']} - {details['pincode']}"
        ]
        
        for i, value in enumerate(company_rows, start=3):
            cell = self.ws[f'A{i}']
            cell.value = value
            cell.font = normal_font
            cell.alignment = left_align

        # Add PAN and Phone in unmerged cells
        pan_cell = self.ws['D3']
        pan_cell.value = f"PAN Number: {details['pan']}"
        pan_cell.font = normal_font
        pan_cell.alignment = left_align

        phone_cell = self.ws['D4']
        phone_cell.value = f"Phone Number: {details['phone']}"
        phone_cell.font = normal_font
        phone_cell.alignment = left_align

    def add_bill_to_section(self, client_details):
        self.ws['A8'].value = "BILL TO"
        self.ws['A8'].fill = subheader_fill
        self.ws['A8'].font = bold_font
        
        self.ws.merge_cells('A8:D8')
        
        start_row = 9
        for key, value in client_details.items():
            self.ws[f'A{start_row}'].value = value
            self.ws.merge_cells(f'A{start_row}:D{start_row}')
            start_row += 1

    def add_items_section(self, items):
        start_row = 15
        
        # Headers
        headers = ['DESCRIPTION', '%tage', '', 'AMOUNT']
        for col, header in enumerate(headers, start=1):
            cell = self.ws[f'{get_column_letter(col)}{start_row}']
            cell.value = header
            cell.fill = subheader_fill
            cell.font = bold_font
            cell.border = border

        # Items
        for item in items:
            start_row += 1
            self.ws[f'A{start_row}'].value = item['description']
            if item.get('percentage'):
                self.ws[f'B{start_row}'].value = item['percentage']
            self.ws[f'D{start_row}'].value = item['amount']
            
        # Totals section
        total_row = start_row + 2
        self.ws[f'C{total_row}'].value = "Payable amount ($)"
        self.ws[f'D{total_row}'].value = "=SUM(D16:D{})".format(start_row)
        
        inr_row = total_row + 1
        self.ws[f'C{inr_row}'].value = "Payable amount (INR)"
        self.ws[f'D{inr_row}'].value = "=D{}*84.14".format(total_row)
        
        value_row = inr_row + 1
        self.ws[f'C{value_row}'].value = "Invoice Value"
        self.ws[f'D{value_row}'].value = f"=D{inr_row}"

    def add_bank_details(self, bank_details):
        start_row = 25
        
        # Apply fill to individual cells first
        for col in ['A', 'B']:
            cell = self.ws[f'{col}{start_row}']
            cell.fill = subheader_fill
        
        # Then merge the cells
        self.ws.merge_cells('A{}:B{}'.format(start_row, start_row))
        
        # Set the header value and font
        header_cell = self.ws[f'A{start_row}']
        header_cell.value = "Bank Account Details"
        header_cell.font = bold_font
        header_cell.alignment = left_align
        
        # Add bank details
        for i, (key, value) in enumerate(bank_details.items(), start=1):
            row = start_row + i
            self.ws[f'A{row}'].value = key
            self.ws[f'B{row}'].value = value
            self.ws[f'A{row}'].font = normal_font
            self.ws[f'B{row}'].font = normal_font
            self.ws[f'A{row}'].alignment = left_align
            self.ws[f'B{row}'].alignment = left_align

    def save(self, filename):
        self.wb.save(filename)
